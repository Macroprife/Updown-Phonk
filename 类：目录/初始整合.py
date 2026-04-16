import os
import re
import pandas as pd
import warnings
from typing import List, Dict, Optional, Union
import xlrd  # 用于读取.xls文件
from openpyxl import load_workbook  # 用于读取.xlsx文件

warnings.filterwarnings('ignore')

def find_excel_files(root_folder: str) -> List[str]:
    """查找指定文件夹及其子文件夹中的所有Excel文件"""
    excel_files = []
    for root, dirs, files in os.walk(root_folder):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(root, file))
    return excel_files

def get_excel_engine(file_path: str) -> str:
    """根据文件扩展名确定使用的引擎"""
    if file_path.lower().endswith('.xlsx'):
        return 'openpyxl'
    elif file_path.lower().endswith('.xls'):
        return 'xlrd'
    else:
        return 'openpyxl'  # 默认

def get_sheet_names(file_path: str) -> List[str]:
    """获取Excel文件中的所有工作表名称"""
    engine = get_excel_engine(file_path)
    
    try:
        if engine == 'openpyxl':
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:  # xlrd
            wb = xlrd.open_workbook(file_path, on_demand=True)
            sheet_names = wb.sheet_names()
            wb.release_resources()
        return sheet_names
    except Exception as e:
        print(f"读取文件 {file_path} 的工作表名称时出错: {str(e)}")
        return []

def extract_max_two_digits_from_a4(file_path: str) -> Optional[int]:
    """从'4.卷内备考表'工作表的A4单元格提取最大的两位数字"""
    try:
        engine = get_excel_engine(file_path)
        
        if engine == 'openpyxl':
            # 使用openpyxl处理.xlsx文件
            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            # 检查是否存在"4.卷内备考表"工作表
            if '4.卷内备考表' not in wb.sheetnames:
                wb.close()
                return None
                
            ws = wb['4.卷内备考表']
            a4_value = ws['A4'].value
            wb.close()
            
        else:  # xlrd
            # 使用xlrd处理.xls文件
            wb = xlrd.open_workbook(file_path)
            
            # 检查是否存在"4.卷内备考表"工作表
            try:
                ws = wb.sheet_by_name('4.卷内备考表')
            except xlrd.biffh.XLRDError:
                wb.release_resources()
                return None
                
            # xlrd的行列索引从0开始
            a4_value = ws.cell_value(3, 0)  # 第4行，第1列 (A4)
            wb.release_resources()
        
        if a4_value is None:
            return None
            
        # 查找所有的两位数字
        two_digit_numbers = re.findall(r'\b\d{2}\b', str(a4_value))
        
        if not two_digit_numbers:
            return None
            
        # 转换为整数并返回最大值
        numbers = [int(num) for num in two_digit_numbers]
        return max(numbers)
        
    except Exception as e:
        print(f"提取文件 {file_path} 的A4单元格数据时出错: {str(e)}")
        return None

def process_excel_file(file_path: str) -> Optional[pd.DataFrame]:
    """处理单个Excel文件，提取所需数据"""
    try:
        # 检查文件是否存在所需的工作表
        sheet_names = get_sheet_names(file_path)
        
        if '3.卷内目录' not in sheet_names:
            return None
            
        # 根据文件类型确定引擎
        engine = get_excel_engine(file_path)
        
        # 读取"3.卷内目录"工作表
        # 跳过前两行，使用第三行作为列名
        try:
            df = pd.read_excel(
                file_path, 
                sheet_name='3.卷内目录', 
                header=2,  # 使用第三行作为列名
                engine=engine # type: ignore
            ) # type: ignore
        except Exception as e:
            print(f"读取文件 {file_path} 的'3.卷内目录'工作表时出错: {str(e)}")
            return None
        
        if df.empty:
            print(f"文件 {file_path} 的'3.卷内目录'工作表为空")
            return None
            
        # 添加源文件名列
        df['源文件名'] = os.path.basename(file_path)
        # 添加文件完整路径列
        df['文件路径'] = file_path
        
        # 提取"每份页数"
        pages_per_file = extract_max_two_digits_from_a4(file_path)
        df['每份页数'] = pages_per_file if pages_per_file is not None else 0
        
        return df
        
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return None

def merge_excel_files(root_folder: str, output_file: str = '汇总结果.xlsx'):
    """主函数：遍历文件夹，合并所有Excel文件的数据"""
    
    print("开始查找Excel文件...")
    excel_files = find_excel_files(root_folder)
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    all_data = []
    processed_files = 0
    skipped_files = 0
    skipped_file_details = []  # 存储跳过的文件信息
    
    for i, file_path in enumerate(excel_files, 1):
        filename = os.path.basename(file_path)
        print(f"\n处理文件 {i}/{len(excel_files)}")
        print(f"文件名: {filename}")
        print(f"完整路径: {file_path}")
        
        df = process_excel_file(file_path)
        
        if df is not None:
            all_data.append(df)
            processed_files += 1
            print(f"状态: ✓ 成功处理")
            print(f"提取数据: {len(df)} 行")
        else:
            skipped_files += 1
            print(f"状态: ✗ 跳过")
            skipped_file_details.append(file_path)
    
    # 输出跳过的文件详细信息
    if skipped_file_details:
        print("\n" + "="*60)
        print("跳过的文件列表:")
        print("="*60)
        for idx, file_path in enumerate(skipped_file_details, 1):
            print(f"{idx:3d}. {file_path}")
    
    if not all_data:
        print("\n错误: 没有找到包含'3.卷内目录'工作表的文件")
        return None
    
    # 合并所有数据
    print("\n合并数据...")
    merged_df = pd.concat(all_data, ignore_index=True)
    
    # 保存到Excel文件
    print(f"\n保存结果到 {output_file}...")
    merged_df.to_excel(output_file, index=False, engine='openpyxl')
    
    print(f"\n处理完成!")
    print(f"处理了 {processed_files} 个文件")
    print(f"跳过了 {skipped_files} 个文件")
    print(f"总记录数: {len(merged_df)}")
    print(f"输出文件大小: {os.path.getsize(output_file) / 1024:.2f} KB")
    
    return merged_df

def show_summary_statistics(df: pd.DataFrame):
    """显示汇总统计信息"""
    print("\n" + "="*60)
    print("汇总统计信息")
    print("="*60)
    
    # 按源文件统计
    print("\n按文件统计记录数:")
    file_counts = df['源文件名'].value_counts()
    for filename, count in file_counts.head(15).items():  # 显示前15个文件
        # 查找对应的完整路径
        file_paths = df[df['源文件名'] == filename]['文件路径'].unique()
        if len(file_paths) > 0:
            path = file_paths[0]
            print(f"  {filename}: {count:4d} 行")
            print(f"      路径: {path}")
    
    if len(file_counts) > 15:
        print(f"  ... 还有 {len(file_counts) - 15} 个文件")
    
    # "每份页数"统计
    print(f"\n'每份页数'统计:")
    print(f"  非零值数量: {(df['每份页数'] > 0).sum()}")
    print(f"  零值数量: {(df['每份页数'] == 0).sum()}")
    if (df['每份页数'] > 0).any():
        print(f"  最小值: {df[df['每份页数'] > 0]['每份页数'].min()}")
        print(f"  最大值: {df[df['每份页数'] > 0]['每份页数'].max()}")
        print(f"  平均值: {df[df['每份页数'] > 0]['每份页数'].mean():.2f}")
    
    # 列信息
    print(f"\n数据列信息 (共 {len(df.columns)} 列):")
    for i, col in enumerate(df.columns, 1):
        non_null = df[col].notna().sum()
        dtype = df[col].dtype
        print(f"  {i:2d}. {col:20s} | 非空值: {non_null:6d} | 类型: {str(dtype):10s}")

def generate_detailed_report(df: pd.DataFrame, root_folder: str, output_file: str):
    """生成详细处理报告"""
    report_file = '详细处理报告.txt'
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("Excel文件处理详细报告\n")
        f.write("="*80 + "\n\n")
        
        f.write(f"处理文件夹: {root_folder}\n")
        f.write(f"输出文件: {output_file}\n")
        f.write(f"处理时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        f.write("一、文件处理统计\n")
        f.write("-"*40 + "\n")
        total_files = len(df['源文件名'].unique())
        f.write(f"成功处理文件数: {total_files}\n")
        f.write(f"总记录行数: {len(df)}\n\n")
        
        f.write("二、处理的文件列表\n")
        f.write("-"*40 + "\n")
        
        # 按文件分组
        for filename in sorted(df['源文件名'].unique()):
            file_data = df[df['源文件名'] == filename]
            file_path = file_data['文件路径'].iloc[0]
            record_count = len(file_data)
            pages = file_data['每份页数'].iloc[0]
            
            f.write(f"\n文件名: {filename}\n")
            f.write(f"完整路径: {file_path}\n")
            f.write(f"记录数: {record_count}\n")
            f.write(f"每份页数: {pages}\n")
        
        f.write("\n三、数据列信息\n")
        f.write("-"*40 + "\n")
        for i, col in enumerate(df.columns, 1):
            non_null = df[col].notna().sum()
            dtype = df[col].dtype
            f.write(f"{i:2d}. {col:20s} | 非空值: {non_null:6d} | 类型: {str(dtype):10s}\n")
        
        f.write("\n四、'每份页数'分布\n")
        f.write("-"*40 + "\n")
        pages_dist = df['每份页数'].value_counts().sort_index()
        for pages, count in pages_dist.items():
            f.write(f"页数 {pages:3d}: {count:5d} 条记录\n")
        
        f.write("\n" + "="*80 + "\n")
        f.write("报告结束\n")
        f.write("="*80 + "\n")
    
    print(f"\n详细报告已生成: {report_file}")
    return report_file

def main():
    """主程序"""
    print("="*80)
    print("Excel文件数据汇总工具 - 带完整路径信息")
    print("="*80)
    print("支持格式: .xlsx 和 .xls")
    print("查找工作表: '3.卷内目录' (使用第3行作为列名)")
    print("提取数据: '4.卷内备考表' 的A4单元格中最大的两位数字作为'每份页数'")
    print("输出包含: 源文件名 + 完整文件路径 + 每份页数")
    print("="*80)
    
    # 设置文件夹路径
    folder_path = input("\n请输入要遍历的文件夹路径: ").strip()
    
    if not os.path.exists(folder_path):
        print("错误: 文件夹不存在!")
        return
    
    # 显示将要搜索的文件夹信息
    print(f"\n将在以下文件夹中搜索Excel文件:")
    print(f"主目录: {folder_path}")
    
    # 设置输出文件名
    output_file = input("\n请输入输出文件名 (默认: 汇总结果.xlsx): ").strip()
    if not output_file:
        output_file = '汇总结果.xlsx'
    
    # 确保输出文件以.xlsx结尾
    if not output_file.lower().endswith('.xlsx'):
        output_file += '.xlsx'
    
    # 检查输出文件是否已存在
    if os.path.exists(output_file):
        overwrite = input(f"\n文件 '{output_file}' 已存在，是否覆盖? (y/n): ").strip().lower()
        if overwrite != 'y':
            print("操作已取消")
            return
    
    # 执行合并
    print("\n开始处理...")
    print("-"*80)
    result = merge_excel_files(folder_path, output_file)
    
    if result is not None:
        # 显示前几行数据
        print("\n" + "="*80)
        print("汇总数据预览 (前5行):")
        print("="*80)
        
        # 只显示部分列用于预览
        preview_cols = ['源文件名', '文件路径', '每份页数']
        other_cols = [col for col in result.columns if col not in preview_cols] # type: ignore
        # 从其他列中选取前3列（如果有的话）
        if other_cols:
            preview_cols.extend(other_cols[:3])
        
        print(result[preview_cols].head().to_string(index=False))
        
        # 显示汇总统计信息
        show_summary_statistics(result)
        
        # 生成详细处理报告
        report_file = generate_detailed_report(result, folder_path, output_file)
        
        print(f"\n" + "="*80)
        print("处理完成!")
        print(f"汇总数据已保存到: {os.path.abspath(output_file)}")
        print(f"详细报告已保存到: {os.path.abspath(report_file)}")
        print("="*80)

if __name__ == "__main__":
    # 检查必要的库是否已安装
    try:
        import pandas
        import xlrd
        from openpyxl import load_workbook
    except ImportError as e:
        print(f"错误: 缺少必要的库 - {e}")
        print("请使用以下命令安装: pip install pandas xlrd openpyxl")
        exit(1)
    
    main()