"""
合并目录.py 输出 → 案卷级 + 文件级 成品表生成
输入：任意 合并目录.py 输出的 xlsx
输出：两个工作表（案卷级 + 文件级），仿劳动桥模板格式
"""

import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ─── 常量（可在此调整默认值） ───────────────
CONTROL_FLAG  = "控制"
KEEP_FOREVER  = "永久"

# 案卷级表头
ANJUAN_HEADERS = [
    "全宗号", "全宗名称", "目录号", "案卷号", "档号",
    "案卷题名", "保管日期", "起止日期", "密级",
    "控制标识", "件数", "页数", "备注"
]

# 文件级表头
WENJIAN_HEADERS = [
    "全宗号", "全宗名称", "目录号", "案卷号", "卷内顺序号",
    "文件档号", "序号", "文号（合同编码/登记薄代码）",
    "责任者", "题名", "日期", "保管期限", "控制标识",
    "页号", "页数", "案卷档号", "备注"
]


def parse_source_filename(filename: str):
    """
    解析源文件名，兼容两种格式：
      230-QQ0309-0001（潘锡勇）.xls
      230-QQ0309-0001潘锡勇.xls
    返回 (全宗号, 目录号, 案卷号, 户主名)
    """
    m = re.match(r"^(\d+)-([A-Z]+\d+)-(\d+)(.*?)\.xls$", filename)
    if not m:
        return None, None, None, None
    quanzong = int(m.group(1))
    mulu     = m.group(2)
    anjuan   = m.group(3)
    raw_name = m.group(4).strip()
    # 去掉全角/半角括号
    huzhu = re.sub(r"[（）()]", "", raw_name).strip()
    return quanzong, mulu, anjuan, huzhu


def extract_group(filepath: str) -> str:
    """从文件路径提取倒数第二级目录作为组名"""
    parts = filepath.replace("\\", "/").split("/")
    return parts[-2] if len(parts) >= 2 else ""


def fmt_date(val) -> str:
    """日期统一转 8 位数字字符串"""
    if pd.isna(val):
        return ""
    if isinstance(val, (int, float)):
        return str(int(val))
    s = str(val).strip()
    return re.sub(r"\D", "", s)[:8] if re.sub(r"\D", "", s) else s


def safe_int(val, default=0) -> int:
    if pd.isna(val) or val == "" or val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


# ═══════════════════════════════════════════
#  0. 手动输入路径
# ═══════════════════════════════════════════
INPUT = input("请输入原始数据文件路径（如 /home/macro/表格/马田.xlsx）: ").strip().strip('"').strip("'")
if not INPUT or not os.path.exists(INPUT):
    print("❌ 文件不存在！")
    exit(1)

OUTPUT = input("请输入输出文件路径（直接回车 = 同目录下生成 原始文件名_成品.xlsx）: ").strip().strip('"').strip("'")
if not OUTPUT:
    base = os.path.splitext(INPUT)[0]
    OUTPUT = f"{base}_成品.xlsx"

QUANZONG_NAME = input("全宗名称（直接回车 = 凯里市土地确权办）: ").strip()
if not QUANZONG_NAME:
    QUANZONG_NAME = "凯里市土地确权办"

TITLE_PREFIX = input("档案题名前缀（含村名，直接回车 = 凯里市万潮镇马田村）: ").strip()
if not TITLE_PREFIX:
    TITLE_PREFIX = "凯里市万潮镇马田村"

# ═══════════════════════════════════════════
#  1. 读入原始数据
# ═══════════════════════════════════════════
print(f"\n📂 读取: {INPUT}")
df = pd.read_excel(INPUT)
df.columns = df.columns.str.strip()  # 列名可能含空格

# 按源文件名分组
groups: dict[str, list] = {}
for _, row in df.iterrows():
    src = row.get("源文件名", "")
    if pd.isna(src):
        continue
    groups.setdefault(src, []).append(row)


# ═══════════════════════════════════════════
#  2. 构造案卷级数据
# ═══════════════════════════════════════════
anjuan_data = []

for src, rows in groups.items():
    first = rows[0]

    # 解析文件名
    qz, ml, aj, hz = parse_source_filename(str(src))
    if not qz:
        print(f"  ⚠ 跳过无法解析的文件名: {src}")
        continue

    danghao = f"{qz}-{ml}-{aj}"
    group_name = extract_group(str(first.get("文件路径", "")))

    # 案卷题名
    title = f"{TITLE_PREFIX}{group_name}{hz}户土地承包经营权确权登记颁证档案"

    # 起止日期
    min_d = fmt_date(first.get("最小日期"))
    max_d = fmt_date(first.get("最大日期"))
    qizhi = f"{min_d}-{max_d}" if min_d and max_d else ""

    # 件数
    jianshu = safe_int(first.get("文件总数"), 0)

    # 页数：优先用 目录页总数，fallback 到 每份页数（根据备考表）
    yeshu = first.get("目录页总数（根据页数相加）")
    if pd.isna(yeshu) or yeshu == "" or yeshu is None:
        yeshu = first.get("每份页数（根据备考表）")
    yeshu = safe_int(yeshu, 0)

    anjuan_data.append([
        qz, QUANZONG_NAME, ml, aj, danghao,
        title, KEEP_FOREVER, qizhi, None,
        CONTROL_FLAG, jianshu, yeshu, None
    ])


# ═══════════════════════════════════════════
#  3. 构造文件级数据
# ═══════════════════════════════════════════
wenjian_data = []

for src, rows in groups.items():
    first = rows[0]
    qz, ml, aj, hz = parse_source_filename(str(src))
    if not qz:
        continue
    danghao = f"{qz}-{ml}-{aj}"

    for seq, row in enumerate(rows, 1):
        # 页号：去空格
        yehao = str(row.get("页号", "") or "").strip()

        # 日期
        date_str = fmt_date(row.get("日期"))

        # 文号（文件编号列名可能含空格）
        wenhao_col = [c for c in df.columns if "文件" in c and "编号" in c]
        wenhao = row.get(wenhao_col[0]) if wenhao_col else None
        if pd.isna(wenhao):
            wenhao = None

        wenjian_data.append([
            qz, QUANZONG_NAME, ml, aj, f"{seq:03d}",
            f"{danghao}-{seq:03d}",
            safe_int(row.get("序号"), seq),
            wenhao,
            row.get("责任者"),
            row.get("文  件  题  名"),
            date_str,
            None,
            CONTROL_FLAG,
            yehao,
            safe_int(row.get("页数"), 0),
            danghao,
            None
        ])


# ═══════════════════════════════════════════
#  4. 输出 Excel
# ═══════════════════════════════════════════
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    # 案卷级工作表
    ws_name_a = "案卷"
    pd.DataFrame(anjuan_data).to_excel(
        writer, sheet_name=ws_name_a, index=False, header=False, startrow=2
    )
    # 文件级工作表
    ws_name_f = "文件"
    pd.DataFrame(wenjian_data).to_excel(
        writer, sheet_name=ws_name_f, index=False, header=False, startrow=2
    )

# ═══════════════════════════════════════════
#  5. 后处理：标题行 + 表头 + 基础格式
# ═══════════════════════════════════════════
wb = load_workbook(OUTPUT)
bold = Font(bold=True, size=11)

for sn, title_text, headers in [
    ("案卷", "机读目录（案卷级）", ANJUAN_HEADERS),
    ("文件", "机读目录（文件级）", WENJIAN_HEADERS),
]:
    ws = wb[sn]

    # Row 1: 标题
    ws.cell(row=1, column=1, value=title_text)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Row 2: 表头
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 自适应列宽（粗略）
    for ci, h in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=ci).column_letter
        max_len = len(h) * 2  # 中文字符约 2 倍宽
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)

wb.save(OUTPUT)

# ═══════════════════════════════════════════
#  6. 统计
# ═══════════════════════════════════════════
print(f"\n✅ 完成！输出文件：{OUTPUT}")
print(f"   案卷级：{len(anjuan_data)} 条")
print(f"   文件级：{len(wenjian_data)} 条")

# 打印摘要
print("\n📋 案卷摘要：")
for row in anjuan_data:
    print(f"   {row[4]}  |  {row[7]}  |  件{row[10]}  页{row[11]}  |  {row[5][:30]}…")

print("\n📄 文件数：")
for src, rows in groups.items():
    _, _, aj, hz = parse_source_filename(str(src)) or (None,)*4
    print(f"   {aj} ({hz}) → {len(rows)} 份文件")

print()
input("按回车键退出...")
