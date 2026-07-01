"""
媒体文件统计工具
================
分别统计图片和 PDF 文件夹，生成 Excel 报告。
GUI 基于 BaseApp 基类。
"""
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from pathlib import Path

import pandas as pd
import PyPDF2
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook

from updown.gui import BaseApp

IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif",
    ".webp", ".svg", ".ico", ".heic", ".heif", ".raw", ".cr2",
}


class MediaStatisticsApp(BaseApp):
    def __init__(self, root):
        super().__init__(root, title="Correct", geometry="750x600",
                         resizable=(True, True))
        self.include_images = tk.BooleanVar(value=True)
        self.include_pdfs   = tk.BooleanVar(value=True)
        self.image_path     = tk.StringVar()
        self.pdf_path       = tk.StringVar()
        self.output_path    = tk.StringVar(value=os.getcwd())
        self.build_ui()

    # ── UI ──────────────────────────────────────────────────

    def build_ui(self):
        mf = ttk.Frame(self.root, padding="20")
        mf.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        mf.columnconfigure(0, weight=1)

        self.add_title(mf, "媒体文件统计", row=0, column=0, pady=(0, 12))

        ttk.Separator(mf, orient="horizontal").grid(
            row=1, column=0, sticky=(tk.W, tk.E), pady=8)

        # ── 图片 ──
        img_frame = ttk.LabelFrame(mf, text="📁 图片文件夹", padding="10")
        img_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=6)
        img_frame.columnconfigure(1, weight=1)

        self._add_toggle_row(img_frame, 0, self.include_images, self.image_path,
                             label="图片统计")
        self.add_browse_folder_row(img_frame, "路径：", self.image_path,
                                   row=1, column=0, columnspan=3)

        fmt_sample = ", ".join(sorted(list(IMAGE_EXTENSIONS))[:8])
        ttk.Label(img_frame, text=f"支持: {fmt_sample}等",
                  foreground="gray").grid(row=2, column=0, columnspan=3,
                                          sticky=tk.W, pady=(3, 0))

        # ── PDF ──
        pdf_frame = ttk.LabelFrame(mf, text="📄 PDF 文件夹", padding="10")
        pdf_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=6)
        pdf_frame.columnconfigure(1, weight=1)

        self._add_toggle_row(pdf_frame, 0, self.include_pdfs, self.pdf_path,
                             label="PDF 统计")
        self.add_browse_folder_row(pdf_frame, "路径：", self.pdf_path,
                                   row=1, column=0, columnspan=3)

        ttk.Label(pdf_frame, text="统计所有子文件夹中的 PDF 页数和数量",
                  foreground="gray").grid(row=2, column=0, columnspan=3,
                                          sticky=tk.W, pady=(3, 0))

        # ── 输出 ──
        out_frame = ttk.LabelFrame(mf, text="💾 输出设置", padding="10")
        out_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=6)
        out_frame.columnconfigure(1, weight=1)
        self.add_browse_folder_row(out_frame, "保存位置：", self.output_path,
                                   row=0, column=0, columnspan=3)

        # ── 进度 + 按钮 ──
        self.add_progress_bar(mf, row=5, column=0, pady=6)
        btn_frame = ttk.Frame(mf)
        btn_frame.grid(row=6, column=0, pady=6)
        self.add_action_button(btn_frame, text="开始统计",
                               command=self.start_statistics, row=0, column=0)
        ttk.Button(btn_frame, text="清空路径",
                   command=self._clear_paths).grid(row=0, column=1, padx=8)
        ttk.Button(btn_frame, text="退出",
                   command=self.root.quit).grid(row=0, column=2)

        # ── 结果 ──
        res_frame = ttk.LabelFrame(mf, text="统计结果", padding="8")
        res_frame.grid(row=7, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=6)
        res_frame.columnconfigure(0, weight=1)
        res_frame.rowconfigure(0, weight=1)
        mf.rowconfigure(7, weight=1)

        text = tk.Text(res_frame, height=10, wrap=tk.WORD,
                       font=self.FONT_LOG, padx=4, pady=4)
        text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        sb = ttk.Scrollbar(res_frame, orient=tk.VERTICAL, command=text.yview)
        sb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        text.config(yscrollcommand=sb.set)
        # 把结果文本框记下来，用于输出统计报告
        self._result_text = text

    def _add_toggle_row(self, parent, row, var, path_var, label="启用"):
        """自定义 ☑/☐ 切换行"""
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(0, 4))

        chk_label = tk.Label(frame, text="☑", font=("Arial", 14),
                             fg="#366092", cursor="hand2")
        chk_label.pack(side=tk.LEFT, padx=(0, 4))

        text_label = tk.Label(frame, text=f"启用{label}（☑启用 / ☐禁用）",
                              font=self.FONT_DEFAULT, cursor="hand2")
        text_label.pack(side=tk.LEFT)

        def toggle(ev=None):
            var.set(not var.get())
            state = "normal" if var.get() else "disabled"
            path_var.set("" if not var.get() else path_var.get())
            chk_label.config(text="☑" if var.get() else "☐",
                             fg="#366092" if var.get() else "gray")

        chk_label.bind("<Button-1>", toggle)
        text_label.bind("<Button-1>", toggle)

    def _clear_paths(self):
        self.image_path.set("")
        self.pdf_path.set("")
        self._result_text.delete(1.0, tk.END)
        self.log("已清空路径")

    # ── 统计逻辑 ────────────────────────────────────────────

    def start_statistics(self):
        img_p = self.image_path.get().strip() if self.include_images.get() else None
        pdf_p = self.pdf_path.get().strip()   if self.include_pdfs.get() else None

        if not self.include_images.get() and not self.include_pdfs.get():
            return messagebox.showwarning("警告", "请至少选择一项统计！")
        if self.include_images.get() and not img_p:
            return messagebox.showwarning("警告", "请选择图片文件夹！")
        if self.include_pdfs.get() and not pdf_p:
            return messagebox.showwarning("警告", "请选择PDF文件夹！")
        if img_p and not os.path.exists(img_p):
            return messagebox.showerror("错误", f"图片文件夹不存在:\n{img_p}")
        if pdf_p and not os.path.exists(pdf_p):
            return messagebox.showerror("错误", f"PDF文件夹不存在:\n{pdf_p}")

        out = self.output_path.get().strip()
        os.makedirs(out, exist_ok=True)

        self.set_busy(True)
        self._result_text.delete(1.0, tk.END)
        threading.Thread(target=self._run, args=(img_p, pdf_p, out),
                         daemon=True).start()

    def _run(self, img_p, pdf_p, out):
        try:
            img_data = img_stats = None
            pdf_data = pdf_stats = None
            if img_p:
                self.log(f"📁 统计图片: {img_p}")
                img_data, img_stats = self._count_images(img_p)
            if pdf_p:
                self.log(f"📁 统计 PDF: {pdf_p}")
                pdf_data, pdf_stats = self._count_pdfs(pdf_p)

            xlsx = self._gen_excel(out, img_data, pdf_data)
            self._show_result(xlsx, img_stats, pdf_stats)
        except Exception as e:
            self.log(f"❌ {e}", "error")
        finally:
            self.root.after(0, lambda: self.set_busy(False))

    def _count_images(self, root):
        data = []
        folders = set()
        total = total_size = 0

        for dirpath, _, fnames in os.walk(root):
            imgs = [f for f in fnames if Path(f).suffix.lower() in IMAGE_EXTENSIONS]
            if not imgs:
                continue
            rel = os.path.relpath(dirpath, root)
            rel = "根目录" if rel == "." else rel
            folders.add(rel)
            sz = sum(os.path.getsize(os.path.join(dirpath, f)) for f in imgs) / 1e6
            total += len(imgs)
            total_size += sz
            data.append({"文件夹": rel, "完整路径": dirpath,
                         "图片数量": len(imgs), "大小(MB)": round(sz, 2)})

        data.sort(key=lambda r: r["图片数量"], reverse=True)
        if data:
            data.append({"文件夹": "【总计】",
                         "完整路径": f"共 {len(folders)} 个文件夹",
                         "图片数量": total,
                         "大小(MB)": round(total_size, 2)})
        return data, {"count": total, "folders": len(folders), "size": round(total_size, 2)}

    def _count_pdfs(self, root):
        data = []
        folders = set()
        total = pages = total_size = 0

        for dirpath, _, fnames in os.walk(root):
            pdfs = [f for f in fnames if f.lower().endswith(".pdf")]
            if not pdfs:
                continue
            rel = os.path.relpath(dirpath, root)
            rel = "根目录" if rel == "." else rel
            folders.add(rel)
            for fn in pdfs:
                fp = os.path.join(dirpath, fn)
                sz = os.path.getsize(fp) / 1e6
                total_size += sz
                try:
                    r = PyPDF2.PdfReader(fp)
                    p = len(r.pages)
                    pages += p
                except Exception:
                    p = "读取失败"
                total += 1
                data.append({"文件夹": rel, "文件名": fn,
                             "页数": p, "大小(MB)": round(sz, 2), "完整路径": fp})

        data.sort(key=lambda r: (r["文件夹"], r["文件名"]))
        if data:
            data.append({"文件夹": "【总计】", "文件名": f"共 {total} 个PDF",
                         "页数": pages, "大小(MB)": round(total_size, 2),
                         "完整路径": f"扫描范围: {root}"})
        return data, {"count": total, "pages": pages,
                      "folders": len(folders), "size": round(total_size, 2)}

    def _gen_excel(self, out, img_data, pdf_data):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx = os.path.join(out, f"媒体文件统计_{ts}.xlsx")

        with pd.ExcelWriter(xlsx, engine="openpyxl") as w:
            if img_data:
                pd.DataFrame(img_data).to_excel(w, sheet_name="图片统计", index=False)
            if pdf_data:
                pd.DataFrame(pdf_data).to_excel(w, sheet_name="PDF统计", index=False)
            if not img_data and not pdf_data:
                pd.DataFrame({"信息": ["未找到文件"]}).to_excel(w, sheet_name="结果", index=False)

        self._fmt_sheets(xlsx)
        return xlsx

    def _fmt_sheets(self, xlsx):
        try:
            wb = load_workbook(xlsx)
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill(start_color="366092", end_color="366092",
                                   fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")
            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = hdr_align
            wb.save(xlsx)
        except Exception:
            pass
        return xlsx

    def _show_result(self, xlsx, img_stats, pdf_stats):
        t = self._result_text
        t.insert(tk.END, "\n" + "=" * 70 + "\n✅ 统计完成！\n\n")

        if img_stats:
            t.insert(tk.END,
                     f"📁 图片: {img_stats['folders']} 个文件夹, "
                     f"{img_stats['count']:,} 张, {img_stats['size']} MB\n")
        if pdf_stats:
            t.insert(tk.END,
                     f"📄 PDF: {pdf_stats['folders']} 个文件夹, "
                     f"{pdf_stats['count']} 个文件, "
                     f"{pdf_stats['pages']:,} 页, {pdf_stats['size']} MB\n")

        t.insert(tk.END, f"\n💾 报告: {xlsx}\n" + "=" * 70 + "\n")

        if messagebox.askyesno("完成", "是否打开 Excel 文件？"):
            os.startfile(xlsx)


def main():
    root = tk.Tk()
    MediaStatisticsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
