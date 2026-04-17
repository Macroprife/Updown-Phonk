import os
import re
import pandas as pd
import warnings
from typing import List, Optional
from datetime import datetime
import xlrd
from openpyxl import load_workbook

warnings.filterwarnings('ignore')

# ==================== Excel汇总功能 ====================

def find_excel_files(root_folder: str) -> List[str]:
    """查找指定文件夹及其子文件夹中的所有Excel文件"""
    excel_files = []
    for root, dirs, files in os.walk(root_folder):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(root, file))
    return excel_files

def get_excel_engine(file_path: str) -> str:
    """根据文件扩展名确定使用的引擎"""
    if file_path.lower().endswith('.xlsx'):
        return 'openpyxl'
    elif file_path.lower().endswith('.xls'):
        return 'xlrd'
    else:
        return 'openpyxl'

def get_sheet_names(file_path: str) -> List[str]:
    """获取Excel文件中的所有工作表名称"""
    engine = get_excel_engine(file_path)
    
    try:
        if engine == 'openpyxl':
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:
            wb = xlrd.open_workbook(file_path, on_demand=True)
            sheet_names = wb.sheet_names()
            wb.release_resources()
        return sheet_names
    except Exception as e:
        print(f"读取文件 {file_path} 的工作表名称时出错: {str(e)}")
        return []

def extract_max_two_digits_from_a4(file_path: str) -> Optional[int]:
    """从'4.卷内备考表'工作表的A4单元格提取最大的两位数字"""
    try:
        engine = get_excel_engine(file_path)
        
        if engine == 'openpyxl':
            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            if '4.卷内备考表' not in wb.sheetnames:
                wb.close()
                return None
                
            ws = wb['4.卷内备考表']
            a4_value = ws['A4'].value
            wb.close()
            
        else:
            wb = xlrd.open_workbook(file_path)
            
            try:
                ws = wb.sheet_by_name('4.卷内备考表')
            except xlrd.biffh.XLRDError:
                wb.release_resources()
                return None
                
            a4_value = ws.cell_value(3, 0)
            wb.release_resources()
        
        if a4_value is None:
            return None
            
        two_digit_numbers = re.findall(r'\b\d{2}\b', str(a4_value))
        
        if not two_digit_numbers:
            return None
            
        numbers = [int(num) for num in two_digit_numbers]
        return max(numbers)
        
    except Exception as e:
        print(f"提取文件 {file_path} 的A4单元格数据时出错: {str(e)}")
        return None

def find_file_title_column(df: pd.DataFrame) -> Optional[str]:
    """查找'文件题名'列（忽略空格）"""
    for col in df.columns:
        col_clean = str(col).replace(' ', '')
        if col_clean == '文件题名':
            return col
    return None

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """清理DataFrame，删除文件题名为空的行"""
    title_col = find_file_title_column(df)
    
    if title_col is None:
        print(f"  警告: 未找到'文件题名'列，跳过清理")
        return df
    
    original_count = len(df)
    
    df = df.dropna(subset=[title_col])
    df = df[df[title_col].astype(str).str.strip() != '']
    
    cleaned_count = len(df)
    removed_count = original_count - cleaned_count
    
    if removed_count > 0:
        print(f"  清理'文件题名'列: 删除 {removed_count} 行空数据")
    
    return df

def process_excel_file_step1(file_path: str) -> Optional[pd.DataFrame]:
    """第一步：处理单个Excel文件，提取所需数据"""
    try:
        sheet_names = get_sheet_names(file_path)
        
        if '3.卷内目录' not in sheet_names:
            return None
            
        engine = get_excel_engine(file_path)
        
        try:
            df = pd.read_excel(
                file_path, 
                sheet_name='3.卷内目录', 
                header=2,
                engine=engine
            )
        except Exception as e:
            print(f"读取文件 {file_path} 的'3.卷内目录'工作表时出错: {str(e)}")
            return None
        
        if df.empty:
            return None
        
        df = clean_dataframe(df)
        
        if df.empty:
            print(f"  清理后无有效数据，跳过此文件")
            return None
            
        df['源文件名'] = os.path.basename(file_path)
        df['文件路径'] = file_path
        
        pages_per_file = extract_max_two_digits_from_a4(file_path)
        df['每份页数'] = pages_per_file if pages_per_file is not None else 0
        
        return df
        
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return None

def merge_excel_files_step1(root_folder: str) -> Optional[pd.DataFrame]:
    """第一步：遍历文件夹，合并所有Excel文件的数据"""
    
    print("正在查找Excel文件...")
    excel_files = find_excel_files(root_folder)
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    all_data = []
    processed_files = 0
    
    for i, file_path in enumerate(excel_files, 1):
        filename = os.path.basename(file_path)
        print(f"处理文件 {i}/{len(excel_files)}: {filename}")
        
        df = process_excel_file_step1(file_path)
        
        if df is not None:
            all_data.append(df)
            processed_files += 1
    
    if not all_data:
        print("错误: 没有找到包含有效数据的文件")
        return None
    
    print("正在合并数据...")
    merged_df = pd.concat(all_data, ignore_index=True)
    
    print(f"成功处理 {processed_files} 个文件，共 {len(merged_df)} 行记录")
    
    return merged_df


# ==================== 统计分析功能 ====================

def parse_date_for_comparison(date_value):
    """将日期转换为可比较的格式"""
    if pd.isna(date_value):
        return pd.NaT
    
    if isinstance(date_value, str):
        date_str = date_value.strip()
        
        if re.match(r'^\d{8}$', date_str):
            try:
                return pd.to_datetime(date_str, format='%Y%m%d')
            except:
                pass
        
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%d-%m-%Y', '%d/%m/%Y']:
            try:
                return pd.to_datetime(date_str, format=fmt)
            except:
                continue
    
    if isinstance(date_value, (int, float)):
        date_str = str(int(date_value))
        if len(date_str) == 8:
            try:
                return pd.to_datetime(date_str, format='%Y%m%d')
            except:
                pass
    
    try:
        return pd.to_datetime(date_value)
    except:
        return pd.NaT

def format_date_to_8digits(date_value):
    """将日期格式化为8位数字字符串"""
    if pd.isna(date_value):
        return ""
    
    if isinstance(date_value, str):
        date_str = date_value.strip()
        if re.match(r'^\d{8}$', date_str):
            return date_str
    
    if isinstance(date_value, (int, float)):
        date_str = str(int(date_value))
        if len(date_str) == 8:
            return date_str
    
    if isinstance(date_value, (pd.Timestamp, datetime)):
        return date_value.strftime('%Y%m%d')
    
    date_str = str(date_value).strip()
    digits_only = re.sub(r'\D', '', date_str)
    
    if len(digits_only) == 8:
        return digits_only
    
    return date_str

def extract_start_page(page_str):
    """提取起始页号"""
    if pd.isna(page_str) or page_str == '':
        return None
    
    page_str = str(page_str)
    
    if '-' in page_str:
        match = re.match(r'(\d+)-', page_str)
        if match:
            return int(match.group(1))
    else:
        match = re.search(r'\d+', page_str)
        if match:
            return int(match.group())
    
    return None

def extract_page_range(page_str):
    """提取页码范围，返回(起始页, 结束页)"""
    if pd.isna(page_str) or page_str == '':
        return None, None
    
    page_str = str(page_str)
    match = re.match(r'(\d+)-(\d+)', page_str)
    if match:
        return int(match.group(1)), int(match.group(2))
    
    return None, None

def process_statistics_step2(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """第二步：对汇总数据进行统计分析"""
    
    print("正在进行统计分析...")
    
    required_columns = ["源文件名", "日期", "页号"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        print(f"错误：缺少必要的列：{missing_columns}")
        return None
    
    original_dates = df["日期"].copy()
    df["日期_转换"] = df["日期"].apply(parse_date_for_comparison)
    
    result_data = []
    
    for source_name, group_df in df.groupby("源文件名"):
        file_count = len(group_df)
        
        max_date_idx = group_df["日期_转换"].idxmax()
        min_date_idx = group_df["日期_转换"].idxmin()
        
        max_date_original = original_dates.loc[max_date_idx]
        min_date_original = original_dates.loc[min_date_idx]
        
        max_date_str = format_date_to_8digits(max_date_original)
        min_date_str = format_date_to_8digits(min_date_original)
        
        group_df_sorted = group_df.sort_index()
        page_numbers = group_df_sorted["页号"].astype(str).tolist()
        
        page_counts = []
        
        for i in range(len(page_numbers)):
            if i < len(page_numbers) - 1:
                current_page = extract_start_page(page_numbers[i])
                next_page = extract_start_page(page_numbers[i + 1])
                
                if current_page is not None and next_page is not None:
                    page_count = next_page - current_page
                else:
                    page_count = 0
                page_counts.append(page_count)
            else:
                if '-' in page_numbers[i]:
                    start_page, end_page = extract_page_range(page_numbers[i])
                    if start_page is not None and end_page is not None:
                        page_count = end_page - start_page + 1
                        
                        if i > 0:
                            prev_page = extract_start_page(page_numbers[i - 1])
                            if prev_page is not None and start_page is not None:
                                page_counts[i - 1] = start_page - prev_page
                    else:
                        page_count = 1
                else:
                    page_count = 0
                
                page_counts.append(page_count)
        
        for i, (row_idx, row) in enumerate(group_df_sorted.iterrows()):
            result_row = row.drop("日期_转换").to_dict()
            result_row["文件总数"] = file_count
            result_row["最大日期"] = max_date_str
            result_row["最小日期"] = min_date_str
            result_row["页数"] = page_counts[i] if i < len(page_counts) else 0
            result_data.append(result_row)
    
    result_df = pd.DataFrame(result_data)
    
    if "日期" in result_df.columns:
        result_df["日期"] = result_df["日期"].apply(format_date_to_8digits)
    
    return result_df


# ==================== 后处理功能 ====================

def add_folder_name_column(df: pd.DataFrame) -> pd.DataFrame:
    """添加总图片文件夹名列（提取源文件名的前15个字符）"""
    if '源文件名' not in df.columns:
        print("警告：没有找到'源文件名'列，跳过添加文件夹名列")
        return df
    
    df['总图片文件夹名'] = df['源文件名'].astype(str).str[:15]
    print(f"已添加'总图片文件夹名'列（提取源文件名的前15个字符）")
    
    return df

def add_sequence_number_column(df: pd.DataFrame) -> pd.DataFrame:
    """添加新列名列（总图片文件夹名-序号）"""
    if '总图片文件夹名' not in df.columns:
        print("警告：没有找到'总图片文件夹名'列，跳过添加序号列")
        return df
    
    # 按"总图片文件夹名"列分组，为每组内的每一行生成序号后缀
    df['新列名'] = df.groupby('总图片文件夹名').cumcount() + 1
    df['新列名'] = df['总图片文件夹名'] + '-' + df['新列名'].astype(str).str.zfill(3)
    print(f"已添加'新列名'列（格式：文件夹名-三位序号）")
    
    return df


# ==================== 主程序 ====================

def get_output_path() -> str:
    """获取输出文件路径"""
    print("\n请选择输出方式：")
    print("1. 手动输入完整路径")
    print("2. 保存在源文件夹中（自动命名）")
    
    choice = input("请选择 (1/2，默认为2): ").strip()
    
    if choice == '1':
        while True:
            output_path = input("请输入输出文件的完整路径（例如：C:\\文件夹\\结果.xlsx）: ").strip()
            output_path = output_path.strip('"').strip("'")
            
            if not output_path:
                print("路径不能为空，请重新输入")
                continue
            
            # 确保以.xlsx结尾
            if not output_path.lower().endswith('.xlsx'):
                output_path += '.xlsx'
            
            # 检查目录是否存在
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                create_dir = input(f"目录 '{output_dir}' 不存在，是否创建？(y/n): ").strip().lower()
                if create_dir == 'y':
                    try:
                        os.makedirs(output_dir)
                        print(f"已创建目录: {output_dir}")
                    except Exception as e:
                        print(f"创建目录失败: {e}")
                        continue
                else:
                    continue
            
            # 检查文件是否已存在
            if os.path.exists(output_path):
                overwrite = input(f"文件 '{output_path}' 已存在，是否覆盖？(y/n): ").strip().lower()
                if overwrite != 'y':
                    continue
            
            return output_path
    else:
        return "auto"

def main():
    """主程序"""
    print("Excel文件数据汇总与统计分析工具")
    
    folder_path = input("\n请输入要处理的文件夹路径: ").strip()
    folder_path = folder_path.strip('"').strip("'")
    
    if not os.path.exists(folder_path):
        print("错误: 文件夹不存在!")
        return
    
    # 获取输出路径
    output_choice = get_output_path()
    
    # 第一步：数据汇总
    df_demo = merge_excel_files_step1(folder_path)
    
    if df_demo is None:
        return
    
    # 第二步：统计分析
    df_result = process_statistics_step2(df_demo)
    
    if df_result is None:
        return
    
    # 第三步：添加文件夹名列
    df_result = add_folder_name_column(df_result)
    
    # 第四步：添加序号列
    df_result = add_sequence_number_column(df_result)
    
    # 确定输出路径
    if output_choice == "auto":
        output_filename = f"BeforeSplit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(folder_path, output_filename)
    else:
        output_path = output_choice
    
    # 保存结果（工作表命名为"抽象"）
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_result.to_excel(writer, sheet_name='抽象', index=False)
    
    print(f"\n处理完成!")
    print(f"输出文件: {output_path}")
    print(f"工作表名: 抽象")
    print(f"新增列: 总图片文件夹名、新列名")

if __name__ == "__main__":
    try:
        import pandas
        import xlrd
        from openpyxl import load_workbook
    except ImportError as e:
        print(f"错误: 缺少必要的库 - {e}")
        print("请安装: pip install pandas xlrd openpyxl")
        exit(1)
    
    main()
    
    input("\n按回车键退出...")
