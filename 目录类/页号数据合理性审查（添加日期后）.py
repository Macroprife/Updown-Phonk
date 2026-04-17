import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def check_page_number_reasonableness(file_path, sheet_name, output_path):
    """
    检查工作表中页号的合理性，并将包含不合理页号的分组的源文件名标黄
    
    参数:
    file_path: 源文件路径
    sheet_name: 工作表名称（"分"）
    output_path: 输出文件路径
    """
    
    # 读取Excel文件
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # 添加"合理性"列，默认值为"目前合理"
    df['合理性'] = '目前合理'
    
    # 用于记录哪些源文件名分组包含不合理数据
    unreasonable_groups = set()
    
    # 按"源文件名"分组
    for source_file, group in df.groupby('源文件名'):
        # 获取该组的索引
        indices = group.index.tolist()
        
        # 获取页号列的数据
        page_numbers = group['页号'].tolist()
        
        # 标记当前分组是否有不合理数据
        has_unreasonable = False
        
        # 检查递进状态（除了最后一行）
        for i in range(len(page_numbers) - 1):
            current_page = page_numbers[i]
            next_page = page_numbers[i+1]
            
            # 跳过最后一行，因为最后一行是特殊格式
            if i == len(page_numbers) - 2:
                continue
                
            # 检查是否为数字且是否递增
            try:
                current_num = float(current_page) if isinstance(current_page, (int, float)) else int(current_page)
                next_num = float(next_page) if isinstance(next_page, (int, float)) else int(next_page)
                
                # 如果不是递增状态，标记为不合理
                if current_num >= next_num:
                    df.loc[indices[i], '合理性'] = '不合理'
                    has_unreasonable = True
                    print(f"警告：在分组 '{source_file}' 中，页号 {current_page} 到 {next_page} 不是递增状态")
                    
            except (ValueError, TypeError):
                # 如果当前行不是最后一行的特殊格式却无法转换为数字，标记为不合理
                df.loc[indices[i], '合理性'] = '不合理'
                has_unreasonable = True
                print(f"警告：在分组 '{source_file}' 中，页号 '{current_page}' 无法识别为有效数字")
        
        # 检查最后一行格式
        if len(page_numbers) > 0:
            last_page = str(page_numbers[-1])
            last_index = indices[-1]
            
            # 检查格式是否为"数字-数字"
            pattern = r'^\d+-\d+$'
            if not re.match(pattern, last_page):
                df.loc[last_index, '合理性'] = '不合理'
                has_unreasonable = True
                print(f"警告：在分组 '{source_file}' 中，最后一行页号 '{last_page}' 不符合XX-XX格式")
            else:
                # 格式正确，检查前后数字（虽然允许一致，但记录一下）
                parts = last_page.split('-')
                if int(parts[0]) > int(parts[1]):
                    # 如果前面的数字大于后面的数字，可能也有问题
                    df.loc[last_index, '合理性'] = '不合理'
                    has_unreasonable = True
                    print(f"警告：在分组 '{source_file}' 中，最后一行页号 '{last_page}' 中前一个数字大于后一个数字")
        
        # 检查倒数第二行到最后一行的递进关系
        if len(page_numbers) >= 2:
            second_last_page = page_numbers[-2]
            last_page_str = str(page_numbers[-1])
            
            try:
                second_last_num = float(second_last_page) if isinstance(second_last_page, (int, float)) else int(second_last_page)
                
                # 提取最后一行的第一个数字
                if re.match(r'^\d+-\d+$', last_page_str):
                    first_num = int(last_page_str.split('-')[0])
                    
                    # 倒数第二行的页号应该小于等于最后一行的第一个数字
                    if second_last_num > first_num:
                        df.loc[indices[-2], '合理性'] = '不合理'
                        has_unreasonable = True
                        print(f"警告：在分组 '{source_file}' 中，倒数第二行页号 {second_last_page} 大于最后一行的起始页 {first_num}")
            except (ValueError, TypeError):
                pass
        
        # 如果该分组包含不合理数据，将源文件名添加到集合中
        if has_unreasonable:
            unreasonable_groups.add(source_file)
    
    # 保存DataFrame到Excel
    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 先保存数据到Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # 使用openpyxl加载工作簿并应用样式
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 创建黄色填充样式
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # 找到"源文件名"列的索引
    source_file_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):  # 第一行是表头
        if cell.value == '源文件名':
            source_file_col_idx = idx
            break
    
    if source_file_col_idx:
        # 遍历数据行（从第2行开始，因为第1行是表头）
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=source_file_col_idx)
            if cell.value in unreasonable_groups:
                cell.fill = yellow_fill
                print(f"标黄处理：分组 '{cell.value}' 的源文件名")
    
    # 保存带格式的工作簿
    wb.save(output_path)
    
    print(f"\n处理完成！结果已保存到: {output_path}")
    
    # 统计不合理记录数
    unreasonable_count = (df['合理性'] == '不合理').sum()
    total_count = len(df)
    unreasonable_group_count = len(unreasonable_groups)
    total_group_count = df['源文件名'].nunique()
    
    print(f"总计 {total_count} 行数据，分布在 {total_group_count} 个分组中")
    print(f"其中 {unreasonable_count} 行被标记为'不合理'")
    print(f"涉及 {unreasonable_group_count} 个不合理的分组，其'源文件名'已标黄处理")
    
    return df

def main():
    """主函数：手动输入路径"""
    print("=" * 50)
    print("页号合理性检查工具（含标黄功能）")
    print("=" * 50)
    
    # 输入源文件路径
    while True:
        file_path = input("\n请输入源Excel文件路径: ").strip().strip('"').strip("'")
        if os.path.exists(file_path):
            break
        else:
            print(f"错误：文件 '{file_path}' 不存在，请重新输入")
    
    # 输入工作表名称
    sheet_name = input("请输入工作表名称（默认为'分'）: ").strip()
    if not sheet_name:
        sheet_name = "分"
    
    # 验证工作表是否存在
    try:
        xl = pd.ExcelFile(file_path)
        if sheet_name not in xl.sheet_names:
            print(f"警告：工作表 '{sheet_name}' 不存在")
            print(f"可用的工作表: {xl.sheet_names}")
            sheet_name = input("请重新输入工作表名称: ").strip()
            if not sheet_name:
                return
    except Exception as e:
        print(f"读取文件出错: {e}")
        return
    
    # 输入输出文件路径
    default_output = os.path.splitext(file_path)[0] + "_合理性检查结果.xlsx"
    output_path = input(f"请输入输出文件路径（默认为'{default_output}'）: ").strip().strip('"').strip("'")
    if not output_path:
        output_path = default_output
    
    # 执行检查
    try:
        check_page_number_reasonableness(file_path, sheet_name, output_path)
        print("\n✅ 检查完成！")
    except Exception as e:
        print(f"\n❌ 处理过程中出错: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # 检查是否安装了openpyxl
    try:
        import openpyxl
    except ImportError:
        print("错误：未安装openpyxl库")
        print("请运行以下命令安装：pip install openpyxl")
        exit(1)
    
    main()
