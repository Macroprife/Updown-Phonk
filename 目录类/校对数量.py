import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import os

def add_group_sum_formulas_new_column(file_path):
    """
    在指定Excel文件的"抽象"工作表中，按源文件名分组，在每组最后一行的新列中添加求和公式
    
    Args:
        file_path: Excel文件路径
    """
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误：文件 {file_path} 不存在")
        return
    
    try:
        # 使用openpyxl加载工作簿以保持公式
        wb = openpyxl.load_workbook(file_path)
        
        # 检查"抽象"工作表是否存在
        if "抽象" not in wb.sheetnames:
            print(f"错误：工作簿中没有名为'抽象'的工作表")
            print(f"可用的工作表：{wb.sheetnames}")
            return
        
        ws = wb["抽象"]
        
        # 使用pandas读取数据以便分析
        df = pd.read_excel(file_path, sheet_name="抽象")
        
        # 检查必需的列是否存在
        required_columns = ["源文件名", "页数"]
        for col in required_columns:
            if col not in df.columns:
                print(f"错误：工作表中没有'{col}'列")
                print(f"可用的列：{df.columns.tolist()}")
                return
        
        # 找到现有列的信息
        headers = [cell.value for cell in ws[1]]
        source_col_idx = headers.index("源文件名") + 1
        pages_col_idx = headers.index("页数") + 1
        pages_col_letter = get_column_letter(pages_col_idx)
        
        print(f"找到'源文件名'列：第{source_col_idx}列")
        print(f"找到'页数'列：第{pages_col_idx}列（{pages_col_letter}列）")
        
        # 确定新列的位置（在现有数据的最后一列之后）
        # 计算数据区域的最大列数
        max_col = ws.max_column
        new_col_idx = max_col + 1
        new_col_letter = get_column_letter(new_col_idx)
        
        # 设置新列的表头
        new_column_name = "分组页数求和"
        ws.cell(row=1, column=new_col_idx, value=new_column_name)
        
        # 设置表头样式
        header_cell = ws.cell(row=1, column=new_col_idx)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        
        print(f"\n将在第{new_col_idx}列（{new_col_letter}列）添加'{new_column_name}'")
        
        # 按源文件名分组
        groups = df.groupby("源文件名")
        
        # 黄色填充样式（用于标记添加公式的单元格）
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 处理每个分组
        formula_count = 0
        for group_name, group_indices in groups.groups.items():
            # 获取该组在Excel中的行号（注意Excel从第2行开始是数据，第1行是表头）
            group_rows = [idx + 2 for idx in group_indices]
            
            if len(group_rows) > 0:
                # 最后一行的Excel行号
                last_row = max(group_rows)
                
                # 第一行的Excel行号
                first_row = min(group_rows)
                
                # 构建求和公式（对"页数"列进行求和）
                formula = f"=SUM({pages_col_letter}{first_row}:{pages_col_letter}{last_row})"
                
                # 在新列的最后一行的对应单元格中设置公式
                cell = ws.cell(row=last_row, column=new_col_idx)
                cell.value = formula
                
                # 用黄色标记该单元格，便于识别
                cell.fill = yellow_fill
                
                # 设置单元格对齐方式
                cell.alignment = Alignment(horizontal='right')
                
                formula_count += 1
                print(f"分组 '{group_name}'：在第{last_row}行{new_col_letter}列添加公式：{formula}")
                
                # 可选：在同一分组的其他行留空或添加注释
                for row in group_rows[:-1]:  # 除了最后一行
                    # 可以选择留空，或者添加注释说明
                    # 这里我们留空，保持工作表整洁
                    pass
        
        # 调整新列的列宽
        ws.column_dimensions[new_col_letter].width = 15
        
        # 保存文件
        output_path = file_path.replace(".xlsx", "_with_group_sum.xlsx").replace(".xls", "_with_group_sum.xls")
        wb.save(output_path)
        
        print(f"\n✅ 成功！")
        print(f"   - 新增了 '{new_column_name}' 列（第{new_col_letter}列）")
        print(f"   - 共添加了 {formula_count} 个求和公式")
        print(f"   - 公式单元格已用黄色标记")
        print(f"   - 文件已保存到：{output_path}")
        
    except Exception as e:
        print(f"处理文件时出错：{str(e)}")
        import traceback
        traceback.print_exc()

# 主程序
if __name__ == "__main__":
    print("=" * 50)
    print("Excel分组求和公式添加工具")
    print("=" * 50)
    print("\n功能说明：")
    print("1. 读取Excel文件的'抽象'工作表")
    print("2. 按'源文件名'列进行分组")
    print("3. 在每组最后一行新增的列中插入求和公式")
    print("4. 原数据保持不变，公式放在新列中")
    print("=" * 50)
    
    # 手动输入文件路径
    file_path = input("\n请输入Excel文件路径（可直接拖拽文件到此处）：").strip()
    
    # 去除可能的引号
    file_path = file_path.strip('"').strip("'")
    
    print(f"\n正在处理文件：{file_path}")
    add_group_sum_formulas_new_column(file_path)
    
    input("\n按Enter键退出...")
